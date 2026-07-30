"""Document parsers for supported file types (issues #9–#12)."""

from __future__ import annotations

from pathlib import Path

from docuwizard.ingest.segments import TextSegment


class ParseError(Exception):
    """Raised when a document cannot be parsed."""


IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}

SUPPORTED_SUFFIXES = {
    ".txt",
    ".md",
    ".markdown",
    ".pdf",
    ".docx",
    ".xlsx",
    ".xlsm",
    ".hwpx",
    *IMAGE_SUFFIXES,
}


def parse_file(path: Path) -> list[TextSegment]:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".markdown"}:
        return parse_text(path)
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix == ".docx":
        return parse_docx(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(path)
    if suffix == ".hwpx":
        return parse_hwpx(path)
    if suffix in IMAGE_SUFFIXES:
        return parse_image(path)
    raise ParseError(f"지원하지 않는 파일 형식입니다: {suffix or path.name}")


def parse_text(path: Path) -> list[TextSegment]:
    try:
        text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = path.read_text(encoding="utf-8", errors="replace")
    segments: list[TextSegment] = []
    for idx, line in enumerate(text.splitlines(), start=1):
        stripped = line.strip()
        if not stripped:
            continue
        segments.append(TextSegment(text=stripped, line_start=idx, line_end=idx))
    if not segments and text.strip():
        segments.append(TextSegment(text=text.strip(), line_start=1, line_end=1))
    return segments


def parse_pdf(path: Path) -> list[TextSegment]:
    from pypdf import PdfReader

    try:
        reader = PdfReader(str(path))
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f"PDF를 열 수 없습니다: {exc}") from exc

    segments: list[TextSegment] = []
    for page_num, page in enumerate(reader.pages, start=1):
        try:
            raw = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            raise ParseError(f"PDF {page_num}페이지 추출 실패: {exc}") from exc
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        if not lines:
            continue
        for line_idx, line in enumerate(lines, start=1):
            segments.append(
                TextSegment(text=line, page=page_num, line_start=line_idx, line_end=line_idx)
            )
    return segments


def parse_docx(path: Path) -> list[TextSegment]:
    from docx import Document

    try:
        document = Document(str(path))
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f"DOCX를 열 수 없습니다: {exc}") from exc

    segments: list[TextSegment] = []
    line = 0
    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue
        line += 1
        segments.append(TextSegment(text=text, line_start=line, line_end=line))
    return segments


def parse_image(path: Path) -> list[TextSegment]:
    """OCR an image with Tesseract (Korean+English, falls back to default)."""
    import pytesseract
    from PIL import Image, UnidentifiedImageError

    try:
        image = Image.open(str(path))
    except (OSError, UnidentifiedImageError) as exc:
        raise ParseError(f"이미지를 열 수 없습니다: {exc}") from exc

    with image:
        try:
            try:
                text = pytesseract.image_to_string(image, lang="kor+eng")
            except pytesseract.TesseractError:
                # Korean language pack may be missing; retry with defaults.
                text = pytesseract.image_to_string(image)
        except pytesseract.TesseractNotFoundError as exc:
            raise ParseError(
                "Tesseract OCR이 설치되어 있지 않습니다. "
                "Windows: https://github.com/UB-Mannheim/tesseract/wiki 에서 설치 "
                "(한국어 언어팩 포함), macOS: brew install tesseract tesseract-lang, "
                "Linux: apt install tesseract-ocr tesseract-ocr-kor"
            ) from exc

    segments: list[TextSegment] = []
    for idx, line in enumerate(text.splitlines(), start=1):
        stripped = line.strip()
        if not stripped:
            continue
        segments.append(TextSegment(text=stripped, line_start=idx, line_end=idx))
    return segments


def parse_hwpx(path: Path) -> list[TextSegment]:
    """Parse HWPX (Hancom Office OWPML) — a ZIP of section XML files.

    Paragraphs are <hp:p> elements; their text lives in <hp:t> children.
    Each paragraph becomes one segment with a sequential line number,
    numbered per section (page metadata uses the section index).
    """
    import re
    import xml.etree.ElementTree as ET
    import zipfile

    try:
        archive = zipfile.ZipFile(str(path))
    except (OSError, zipfile.BadZipFile) as exc:
        raise ParseError(f"HWPX를 열 수 없습니다: {exc}") from exc

    def section_index(name: str) -> int:
        match = re.search(r"section(\d+)\.xml$", name)
        return int(match.group(1)) if match else 0

    segments: list[TextSegment] = []
    with archive:
        section_names = sorted(
            (
                name
                for name in archive.namelist()
                if name.startswith("Contents/section") and name.endswith(".xml")
            ),
            key=section_index,
        )
        if not section_names:
            raise ParseError("HWPX에서 본문(section)을 찾을 수 없습니다.")
        for section_no, name in enumerate(section_names, start=1):
            try:
                root = ET.fromstring(archive.read(name))
            except ET.ParseError as exc:
                raise ParseError(f"HWPX 본문 파싱 실패({name}): {exc}") from exc
            line = 0
            for paragraph in root.iter():
                if not paragraph.tag.endswith("}p"):
                    continue
                texts = [
                    el.text
                    for el in paragraph.iter()
                    if el.tag.endswith("}t") and el.text
                ]
                text = "".join(texts).strip()
                if not text:
                    continue
                line += 1
                segments.append(
                    TextSegment(
                        text=text,
                        page=section_no if len(section_names) > 1 else None,
                        line_start=line,
                        line_end=line,
                    )
                )
    return segments


def parse_xlsx(path: Path) -> list[TextSegment]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f"Excel 파일을 열 수 없습니다: {exc}") from exc

    segments: list[TextSegment] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                cells = [cell for cell in row if cell.value is not None and str(cell.value).strip()]
                if not cells:
                    continue
                parts = [str(cell.value).strip() for cell in cells]
                first = cells[0].coordinate
                last = cells[-1].coordinate
                cell_range = first if first == last else f"{first}:{last}"
                segments.append(
                    TextSegment(
                        text=" | ".join(parts),
                        sheet=sheet.title,
                        cell_range=cell_range,
                    )
                )
    finally:
        workbook.close()
    return segments
